"""
Command-line interface for AI Agent Directory scraper
"""
import json
import typer
from src.services.scraper import AIDirectoryScraper
from wordcloud import WordCloud
import matplotlib.pyplot as plt
from pathlib import Path

app = typer.Typer(pretty_exceptions_show_locals=False)

@app.command()
def scrape():
    with open('./update.json', 'r') as fd:
        agent_map = json.load(fd)

    """Scrape AI Agent Directory"""
    scraper = AIDirectoryScraper(headless=True)
    try:
        categories = scraper.get_categories()
        print(f"Found {len(categories)} categories")
        
        all_agents = []
        for category_name, category_url in categories.items():
            category_agents = scraper.get_agents_from_category(category_url)
            print(f"Found {len(category_agents)} agents in category {category_name}")
            for agent in category_agents: 
                if not agent.name in agent_map or None == agent_map[agent.name].get('info'):
                    agent_with_details = scraper.get_agent_details(agent)
                    all_agents.append(agent_with_details)
                    scraper.save_progress(all_agents)
                else:
                    all_agents.append(agent_map[agent.name])
                    scraper.save_progress(all_agents)

        print(f"Processed {len(all_agents)} agents with details")
    finally:
        scraper.close()


@app.command()
def version():
    """Show version information"""
    print("AI Agent Directory Scraper v0.1.0")


@app.command()
def transform_to_excel(category: str=None):
    """Transform JSON data to Excel"""
    with open('./ai_agents_data.json', 'r') as fd:
        data = json.load(fd)

    if category:
        agent_list = [agent for agent in data if agent['source_url'].endswith(category)]
    else:
        agent_list = [agent for agent in data]

    # create a excel file with openpyxl
    from openpyxl import Workbook
    from openpyxl.styles import Alignment
    wb = Workbook()
    ws = wb.active
    if category:
        ws.title = category
    else:
        ws.title = 'all'


    ws['A1'] = 'Name'
    ws['B1'] = 'Review'
    ws['C1'] = 'Url'
    ws['D1'] = 'Category'
    ws['E1'] = 'Key Features'
    ws['F1'] = 'User Cases'
    ws['G1'] = 'Details'
    for idx, agent in enumerate(agent_list): 
        ws[f'A{idx+2}'] = agent['name']
        print(agent['name'])
        ws[f'B{idx+2}'] = agent['info']['review']
        ws[f'C{idx+2}'] = agent['url']
        ws[f'D{idx+2}'] = agent['source_url'].split('/')[-1]
        ws[f'E{idx+2}'].alignment = Alignment(wrap_text=True)
        ws[f'E{idx+2}'] = '\n'.join(agent['info']['key_features'])
        ws[f'F{idx+2}'].alignment = Alignment(wrap_text=True)
        ws[f'F{idx+2}'] = '\n'.join(agent['info']['user_cases'])
        ws[f'G{idx+2}'].alignment = Alignment(wrap_text=True)
        ws[f'G{idx+2}'] = '\n'.join([f"{k}: {v}" for k, v in agent['info']['details'].items()]) 
    wb.save('ai_agents_data.xlsx')


@app.command()
def generate_wordcloud():
    """Generate a Minecraft-style word cloud from user cases in AI agents data"""
    # Load the data
    try:
        with open('./ai_agents_data.json', 'r') as fd:
            data = json.load(fd)
    except FileNotFoundError:
        print("Error: ai_agents_data.json not found")
        return
    except json.JSONDecodeError:
        print("Error: Invalid JSON in ai_agents_data.json")
        return
    
    # Extract all user cases
    all_user_cases = []
    for agent in data:
        if isinstance(agent, dict) and isinstance(agent.get('info'), dict):
            user_cases = agent['info'].get('user_cases', [])
            if isinstance(user_cases, list):
                if 'User Case 1' in user_cases:
                    continue
                all_user_cases.extend(user_cases)
    
    if not all_user_cases:
        print("No user cases found in the data")
        return
    
    # Join all user cases into a single string
    text = ' '.join(all_user_cases)
    
    # Create and configure the WordCloud object
    wordcloud = WordCloud(
        width=800,
        height=400,
        background_color='#C6C6C6',  # Minecraft stone-like gray
        colormap='YlOrRd',  # Minecraft-like colors (yellows, oranges, reds)
        prefer_horizontal=0.7,
        font_path='/System/Library/Fonts/Supplemental/Minecraft.ttf' if Path('/System/Library/Fonts/Supplemental/Minecraft.ttf').exists() else None,
        min_font_size=10,
        max_font_size=60,
        relative_scaling=0.5
    ).generate(text)
    
    # Create the plot
    plt.figure(figsize=(16, 8))
    plt.imshow(wordcloud, interpolation='bilinear')
    plt.axis('off')
    
    # Save the word cloud
    output_path = 'minecraft_wordcloud.png'
    plt.savefig(output_path, bbox_inches='tight', pad_inches=0)
    plt.close()
    
    print(f"Word cloud saved as {output_path}")


if __name__ == "__main__":
    app()
